import pandas as pd
import os
from openpyxl import load_workbook

def verify_no_merged_cells(file_path):
    """验证Excel文件中没有合并单元格"""
    try:
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            merged_ranges = list(worksheet.merged_cells.ranges)
            if merged_ranges:
                print(f"⚠️  工作表 '{sheet_name}' 中发现 {len(merged_ranges)} 个合并单元格:")
                for merged_range in merged_ranges:
                    print(f"   - {merged_range}")
                return False
            else:
                print(f"✅ 工作表 '{sheet_name}' 中没有合并单元格")
        return True
    except Exception as e:
        print(f"❌ 验证文件时出错: {e}")
        return False

def compare_dimensions(original_file, transposed_file):
    """比较原始文件和转置文件的维度"""
    try:
        # 读取原始文件
        original_df = pd.read_excel(original_file, header=None)
        original_cleaned = original_df.dropna(how='all').dropna(axis=1, how='all')
        
        # 读取转置文件
        transposed_df = pd.read_excel(transposed_file, header=None)
        
        print(f"\n📊 维度对比:")
        print(f"原始文件: {original_cleaned.shape[0]}行 × {original_cleaned.shape[1]}列")
        print(f"转置文件: {transposed_df.shape[0]}行 × {transposed_df.shape[1]}列")
        
        # 检查转置是否正确（行列互换）
        if (original_cleaned.shape[0] == transposed_df.shape[1] and 
            original_cleaned.shape[1] == transposed_df.shape[0]):
            print("✅ 转置维度正确")
            return True
        else:
            print("❌ 转置维度不正确")
            return False
            
    except Exception as e:
        print(f"❌ 比较维度时出错: {e}")
        return False

def display_file_preview(file_path, title):
    """显示文件预览"""
    try:
        df = pd.read_excel(file_path, header=None)
        print(f"\n{title} (前5行5列):")
        preview = df.iloc[:5, :5]
        print(preview.to_string())
    except Exception as e:
        print(f"❌ 读取 {title} 时出错: {e}")

def main():
    print("=== 转置结果验证程序 ===\n")
    
    # 查找输出文件
    output_dir = "outputs"
    if not os.path.exists(output_dir):
        print(f"❌ 输出目录不存在: {output_dir}")
        return
    
    # 找到最新的转置文件
    output_files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx') and '转置' in f]
    if not output_files:
        print("❌ 未找到转置文件")
        return
    
    # 按修改时间排序，获取最新文件
    output_files.sort(key=lambda x: os.path.getmtime(os.path.join(output_dir, x)), reverse=True)
    latest_file = os.path.join(output_dir, output_files[0])
    
    print(f"📁 验证文件: {latest_file}")
    
    # 原始文件路径
    original_file = "待处理文件/2025926移山科技循环10次采集对内报表_副本.xlsx"
    
    # 1. 验证没有合并单元格
    print("\n1️⃣ 检查合并单元格...")
    no_merged = verify_no_merged_cells(latest_file)
    
    # 2. 比较维度
    print("\n2️⃣ 验证转置维度...")
    correct_dimensions = compare_dimensions(original_file, latest_file)
    
    # 3. 显示文件预览
    print("\n3️⃣ 文件内容预览...")
    display_file_preview(original_file, "原始文件")
    display_file_preview(latest_file, "转置文件")
    
    # 总结
    print(f"\n📋 验证结果:")
    print(f"   ✅ 无合并单元格: {'是' if no_merged else '否'}")
    print(f"   ✅ 转置维度正确: {'是' if correct_dimensions else '否'}")
    
    if no_merged and correct_dimensions:
        print(f"\n🎉 验证通过！转置文件符合要求。")
        print(f"📄 输出文件: {latest_file}")
    else:
        print(f"\n⚠️  验证未完全通过，请检查上述问题。")

if __name__ == "__main__":
    main()