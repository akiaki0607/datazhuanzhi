#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
信源数据分析转置处理工具
基于示例文件格式进行转置处理
"""

import pandas as pd
import openpyxl
import os
import sys
from datetime import datetime

def process_source_data_transpose(input_file, output_file=None):
    """
    信源数据分析转置处理函数
    
    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径（可选，默认自动生成）
    
    返回:
    DataFrame: 转换后的数据
    """
    try:
        print(f"开始处理文件: {input_file}")
        
        # 检查输入文件是否存在
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"找不到输入文件: {input_file}")
        
        # 第一步：文件结构分析
        wb = openpyxl.load_workbook(input_file, data_only=True)
        
        # 检查是否有"信源数据分析"工作表
        if "信源数据分析" not in wb.sheetnames:
            raise ValueError("找不到'信源数据分析'工作表")
        
        ws = wb["信源数据分析"]
        
        print(f"工作表: 信源数据分析")
        print(f"工作表尺寸: {ws.max_row} 行 x {ws.max_column} 列")
        
        # 识别合并单元格
        merged_ranges = list(ws.merged_cells.ranges)
        print(f"找到 {len(merged_ranges)} 个合并单元格区域")
        
        # 分析表头结构
        print("分析表头结构...")
        
        # 获取第一行（品牌标题）
        brand_headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            brand_headers.append(cell.value)
        
        # 获取第二行（子标题）
        sub_headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col_idx)
            sub_headers.append(cell.value)
        
        print(f"品牌标题行: {brand_headers[:15]}...")  # 显示前15个
        print(f"子标题行: {sub_headers[:15]}...")  # 显示前15个
        
        # 识别品牌列
        brand_columns = {}
        for merged_range in merged_ranges:
            top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
            brand_name = top_left_cell.value
            if brand_name and isinstance(brand_name, str) and brand_name.strip():
                brand_columns[brand_name] = {
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col
                }
                print(f"  品牌: {brand_name} (列 {merged_range.min_col}-{merged_range.max_col})")
        
        # 如果没有找到合并单元格，尝试从第一行识别品牌
        if not brand_columns:
            print("未找到合并单元格，尝试从第一行识别品牌...")
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                brand_name = cell.value
                if brand_name and isinstance(brand_name, str) and brand_name.strip():
                    # 查找该品牌对应的列范围
                    start_col = col_idx
                    end_col = col_idx
                    # 向后查找连续的空列或相同品牌
                    for next_col in range(col_idx + 1, ws.max_column + 1):
                        next_cell = ws.cell(row=1, column=next_col)
                        if next_cell.value is None or next_cell.value == '':
                            end_col = next_col
                        else:
                            break
                    
                    brand_columns[brand_name] = {
                        'start_col': start_col,
                        'end_col': end_col
                    }
                    print(f"  品牌: {brand_name} (列 {start_col}-{end_col})")
        
        # 第二步：建立品牌到列的映射
        brand_to_columns = {}
        for brand_name, col_info in brand_columns.items():
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            brand_sub_headers = []
            
            for i in range(start_col, end_col + 1):
                if i <= len(sub_headers):
                    brand_sub_headers.append(sub_headers[i-1])
            
            brand_to_columns[brand_name] = {
                'start_col': start_col,
                'end_col': end_col,
                'sub_headers': brand_sub_headers
            }
        
        # 第三步：数据提取
        print("开始提取数据...")
        data_rows = []
        data_start_row = 3  # 数据从第3行开始
        
        for row_idx in range(data_start_row, ws.max_row + 1):
            # 获取关键词名称（第一列）
            keyword = ws.cell(row=row_idx, column=1).value
            # 获取AI平台名称（第二列）
            ai_platform = ws.cell(row=row_idx, column=2).value
            # 获取信源平台名称（第三列）
            source_platform = ws.cell(row=row_idx, column=3).value
            
            if keyword is None or keyword == '':
                continue
            
            # 为每个品牌提取数据
            for brand_name, col_info in brand_to_columns.items():
                row_data = {
                    '关键词名称': keyword,
                    'AI平台名称': ai_platform,
                    '信源平台名称': source_platform,
                    '品牌': brand_name
                }
                
                # 提取该品牌对应的数据列
                start_col = col_info['start_col']
                end_col = col_info['end_col']
                sub_headers = col_info['sub_headers']
                
                for i, sub_header in enumerate(sub_headers):
                    col_idx = start_col + i
                    if col_idx <= end_col and col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        row_data[sub_header] = cell_value
                    else:
                        row_data[sub_header] = None
                
                data_rows.append(row_data)
            
            # 显示进度
            if row_idx % 100 == 0:
                print(f"已处理 {row_idx - data_start_row + 1} 行数据...")
        
        print(f"数据提取完成，总共 {len(data_rows)} 行数据")
        
        # 第四步：创建DataFrame并验证
        df = pd.DataFrame(data_rows)
        print(f"转换后的数据形状: {df.shape}")
        
        # 数据验证
        print("\n数据验证:")
        print(f"总行数: {len(df)}")
        print(f"总列数: {len(df.columns)}")
        print(f"唯一关键词数: {df['关键词名称'].nunique()}")
        print(f"唯一AI平台数: {df['AI平台名称'].nunique()}")
        print(f"唯一信源平台数: {df['信源平台名称'].nunique()}")
        print(f"唯一品牌数: {df['品牌'].nunique()}")
        
        # 检查非空数据
        non_empty_total = 0
        for col in df.columns:
            if col not in ['关键词名称', 'AI平台名称', '信源平台名称', '品牌']:
                non_empty = df[col].notna() & (df[col] != '0.0%') & (df[col] != 0) & (df[col] != '0')
                non_empty_count = non_empty.sum()
                non_empty_total += non_empty_count
                if non_empty_count > 0:
                    print(f"{col}列非空数据: {non_empty_count} 条")
        
        print(f"总非空数据条目: {non_empty_total}")
        
        # 第五步：保存文件
        if output_file is None:
            # 自动生成输出文件名，包含日期和转置完成标记
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            current_date = datetime.now().strftime("%Y%m%d")
            output_file = f"{base_name}_{current_date}_转置完成.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='转置后数据', index=False)
        
        file_size = os.path.getsize(output_file)
        print(f"\n文件已保存: {output_file}")
        print(f"文件大小: {file_size / 1024:.2f} KB")
        
        return df
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """
    主函数 - 命令行使用
    """
    if len(sys.argv) < 2:
        print("使用方法: python process_source_data.py <输入文件> [输出文件]")
        print("示例: python process_source_data.py 数据文件.xlsx")
        print("示例: python process_source_data.py 数据文件.xlsx 输出文件.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    print("=" * 60)
    print("信源数据分析转置处理工具")
    print("=" * 60)
    
    df = process_source_data_transpose(input_file, output_file)
    
    if df is not None:
        print("\n" + "=" * 60)
        print("处理完成!")
        print("=" * 60)
    else:
        print("\n处理失败!")

if __name__ == "__main__":
    main()

