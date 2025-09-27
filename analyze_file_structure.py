#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析Excel文件结构
"""

import openpyxl
import os

def analyze_excel_structure(file_path):
    """
    分析Excel文件结构
    """
    try:
        print(f"分析文件: {file_path}")
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"工作表数量: {len(wb.sheetnames)}")
        print(f"工作表名称: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n=== 工作表: {sheet_name} ===")
            ws = wb[sheet_name]
            print(f"尺寸: {ws.max_row} 行 x {ws.max_column} 列")
            
            # 显示前几行数据
            print("前5行数据:")
            for row_idx in range(1, min(6, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(11, ws.max_column + 1)):  # 只显示前10列
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is None:
                        value = ""
                    row_data.append(str(value)[:20])  # 限制长度
                print(f"  第{row_idx}行: {row_data}")
            
            # 检查合并单元格
            merged_ranges = list(ws.merged_cells.ranges)
            print(f"合并单元格数量: {len(merged_ranges)}")
            if merged_ranges:
                for i, merged_range in enumerate(merged_ranges[:5]):  # 只显示前5个
                    print(f"  合并区域{i+1}: {merged_range}")
        
    except Exception as e:
        print(f"分析过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    file_path = "待处理文件/2025926移山科技循环10次采集对内报表.xlsx"
    analyze_excel_structure(file_path)

