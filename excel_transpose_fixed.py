#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel合并单元格转置工具 - 修正版
正确提取和处理所有数据，确保AI平台数据不丢失
"""

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def process_merged_cells_to_long_format_fixed(file_path, sheet_name=None):
    """
    修正版：将包含合并单元格的Excel数据转换为长格式
    """
    print(f"正在处理Excel文件: {file_path}")
    
    try:
        # 使用openpyxl读取工作簿
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 检查工作表
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]
        
        print(f"使用工作表: {sheet_name}")
        ws = wb[sheet_name]
        
        print(f"工作表尺寸: {ws.max_row} 行 x {ws.max_column} 列")
        
        # 分析合并单元格
        merged_ranges = list(ws.merged_cells.ranges)
        print(f"找到 {len(merged_ranges)} 个合并单元格区域")
        
        # 创建品牌到列索引的映射
        brand_columns = {}
        for merged_range in merged_ranges:
            top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
            brand_name = top_left_cell.value
            
            if brand_name and isinstance(brand_name, str):
                brand_columns[brand_name] = {
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col,
                    'row': merged_range.min_row
                }
        
        print(f"识别到 {len(brand_columns)} 个品牌")
        
        # 获取子标题行（第2行）
        sub_headers_row = 2
        sub_headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=sub_headers_row, column=col_idx)
            sub_headers.append(cell.value)
        
        print(f"子标题: {sub_headers[:10]}...")  # 显示前10个
        
        # 为每个品牌确定其子标题
        brand_to_columns = {}
        for brand_name, col_info in brand_columns.items():
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            
            # 获取该品牌对应的子标题
            brand_sub_headers = []
            for i in range(start_col, end_col + 1):
                if i <= len(sub_headers):
                    brand_sub_headers.append(sub_headers[i-1])
            
            brand_to_columns[brand_name] = {
                'start_col': start_col,
                'end_col': end_col,
                'sub_headers': brand_sub_headers
            }
            
            print(f"品牌: {brand_name}")
            print(f"  列范围: {start_col}-{end_col}")
            print(f"  子标题: {brand_sub_headers}")
        
        # 提取所有数据
        data_rows = []
        data_start_row = 3  # 数据从第3行开始
        
        print(f"\n开始提取数据，从第{data_start_row}行到第{ws.max_row}行...")
        
        for row_idx in range(data_start_row, ws.max_row + 1):
            # 获取信源平台名称（第一列）
            source_platform = ws.cell(row=row_idx, column=1).value
            
            if source_platform is None or source_platform == '':
                continue
            
            # 为每个品牌提取数据
            for brand_name, col_info in brand_to_columns.items():
                row_data = {
                    '信源平台名称': source_platform,
                    '品牌': brand_name
                }
                
                # 提取该品牌对应的数据列
                start_col = col_info['start_col']
                end_col = col_info['end_col']
                sub_headers = col_info['sub_headers']
                
                for i, sub_header in enumerate(sub_headers):
                    col_idx = start_col + i
                    if col_idx <= end_col and col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        row_data[sub_header] = cell_value
                    else:
                        row_data[sub_header] = None
                
                data_rows.append(row_data)
            
            # 显示进度
            if row_idx % 50 == 0:
                print(f"已处理 {row_idx - data_start_row + 1} 行数据...")
        
        print(f"数据提取完成，总共 {len(data_rows)} 行数据")
        
        # 创建DataFrame
        df = pd.DataFrame(data_rows)
        print(f"转换后的数据形状: {df.shape}")
        
        # 检查是否有非空数据
        non_empty_count = 0
        for col in ['DeepSeek', 'Kimi', '元宝', '豆包']:
            if col in df.columns:
                non_empty = df[col].notna() & (df[col] != '0.0%') & (df[col] != 0) & (df[col] != '0')
                non_empty_count += non_empty.sum()
        
        print(f"非空数据条目数量: {non_empty_count}")
        
        # 显示一些示例数据
        print("\n转换后的数据预览:")
        print(df.head(10))
        
        # 显示一些非零数据的示例
        print("\n非零数据示例:")
        for col in ['DeepSeek', 'Kimi', '元宝', '豆包']:
            if col in df.columns:
                non_zero_data = df[df[col].notna() & (df[col] != '0.0%') & (df[col] != 0) & (df[col] != '0')]
                if len(non_zero_data) > 0:
                    print(f"{col}列非零数据示例:")
                    print(non_zero_data[['信源平台名称', '品牌', col]].head(5))
                    break
        
        return df
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def save_processed_data(df, output_file):
    """
    保存处理后的数据到新的Excel文件
    """
    if df is None:
        print("没有数据需要保存")
        return
    
    try:
        # 创建新的Excel文件
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='转置后数据', index=False)
        
        print(f"数据已保存到: {output_file}")
        
        # 显示保存的文件信息
        file_size = os.path.getsize(output_file)
        print(f"文件大小: {file_size / 1024:.2f} KB")
        
        # 统计信息
        print(f"总行数: {len(df)}")
        print(f"总列数: {len(df.columns)}")
        print(f"唯一信源平台数: {df['信源平台名称'].nunique()}")
        print(f"唯一品牌数: {df['品牌'].nunique()}")
        
    except Exception as e:
        print(f"保存文件时出现错误: {str(e)}")
        import traceback
        traceback.print_exc()

def main():
    """
    主函数
    """
    # 输入文件路径
    input_file = "2025916移山科技循环10次采集任务34词对外报表_待处理.xlsx"
    output_file = "2025916移山科技循环10次采集任务34词对外报表_转置后_修正版.xlsx"
    
    print("=" * 60)
    print("Excel合并单元格转置工具 - 修正版")
    print("=" * 60)
    
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 找不到输入文件 {input_file}")
        return
    
    # 处理数据
    df = process_merged_cells_to_long_format_fixed(input_file)
    
    if df is not None:
        # 保存结果
        save_processed_data(df, output_file)
        
        print("\n" + "=" * 60)
        print("处理完成!")
        print("=" * 60)
        print(f"输入文件: {input_file}")
        print(f"输出文件: {output_file}")
        print(f"转换后数据行数: {len(df)}")
        print(f"转换后数据列数: {len(df.columns)}")
        
        # 验证数据完整性
        print("\n数据验证:")
        for col in ['DeepSeek', 'Kimi', '元宝', '豆包']:
            if col in df.columns:
                non_empty = df[col].notna() & (df[col] != '0.0%') & (df[col] != 0) & (df[col] != '0')
                print(f"{col}列非空数据: {non_empty.sum()} 条")
    else:
        print("处理失败，请检查输入文件格式")

if __name__ == "__main__":
    main()

