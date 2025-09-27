#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
三次测试对内报表转置工具
专门处理包含多个工作表的对内报表文件
"""

import pandas as pd
import openpyxl
import os
import sys

def process_three_test_report(input_file, output_file=None):
    """
    处理三次测试对内报表的转置
    
    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径（可选，默认自动生成）
    
    返回:
    dict: 包含各工作表转换后数据的字典
    """
    try:
        print(f"开始处理文件: {input_file}")
        
        # 检查输入文件是否存在
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"找不到输入文件: {input_file}")
        
        # 加载工作簿
        wb = openpyxl.load_workbook(input_file, data_only=True)
        print(f"工作表列表: {wb.sheetnames}")
        
        results = {}
        
        # 处理汇总报表
        if '汇总报表' in wb.sheetnames:
            print("\n处理汇总报表...")
            ws = wb['汇总报表']
            df_summary = pd.DataFrame(ws.values)
            df_summary.columns = df_summary.iloc[0]
            df_summary = df_summary.drop(0).reset_index(drop=True)
            results['汇总报表'] = df_summary
            print(f"汇总报表数据形状: {df_summary.shape}")
        
        # 处理关键词数据分析
        if '关键词数据分析' in wb.sheetnames:
            print("\n处理关键词数据分析...")
            ws = wb['关键词数据分析']
            
            # 找到品牌列的位置
            brand_columns = {}
            for row_idx in range(1, min(6, ws.max_row + 1)):
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value and isinstance(cell_value, str) and cell_value not in ['None', '关键词名称', 'AI平台名称']:
                        # 检查这个值是否在多个列中出现（合并单元格的特征）
                        brand_name = cell_value
                        if brand_name not in brand_columns:
                            brand_columns[brand_name] = []
                        brand_columns[brand_name].append(col_idx)
            
            print(f"找到品牌: {list(brand_columns.keys())}")
            
            # 获取子标题（第2行）
            sub_headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=2, column=col_idx)
                sub_headers.append(cell.value)
            
            # 建立品牌到列的映射
            brand_to_columns = {}
            for brand_name, col_indices in brand_columns.items():
                if len(col_indices) > 1:  # 合并单元格
                    start_col = min(col_indices)
                    end_col = max(col_indices)
                    brand_sub_headers = sub_headers[start_col-1:end_col]
                else:  # 单列
                    start_col = col_indices[0]
                    end_col = col_indices[0]
                    brand_sub_headers = [sub_headers[start_col-1]] if start_col <= len(sub_headers) else []
                
                brand_to_columns[brand_name] = {
                    'start_col': start_col,
                    'end_col': end_col,
                    'sub_headers': brand_sub_headers
                }
            
            # 提取数据
            data_rows = []
            data_start_row = 3  # 数据从第3行开始
            
            for row_idx in range(data_start_row, ws.max_row + 1):
                # 获取关键词名称（第一列）
                keyword_name = ws.cell(row=row_idx, column=1).value
                
                if keyword_name is None or keyword_name == '':
                    continue
                
                # 为每个品牌提取数据
                for brand_name, col_info in brand_to_columns.items():
                    row_data = {
                        '关键词名称': keyword_name,
                        '品牌': brand_name
                    }
                    
                    # 提取该品牌对应的数据列
                    start_col = col_info['start_col']
                    end_col = col_info['end_col']
                    sub_headers = col_info['sub_headers']
                    
                    for i, sub_header in enumerate(sub_headers):
                        col_idx = start_col + i
                        if col_idx <= end_col and col_idx <= ws.max_column:
                            cell_value = ws.cell(row=row_idx, column=col_idx).value
                            row_data[sub_header] = cell_value
                        else:
                            row_data[sub_header] = None
                    
                    data_rows.append(row_data)
            
            df_keywords = pd.DataFrame(data_rows)
            results['关键词数据分析_转置'] = df_keywords
            print(f"关键词数据分析转置后形状: {df_keywords.shape}")
        
        # 处理信源数据分析
        if '信源数据分析' in wb.sheetnames:
            print("\n处理信源数据分析...")
            ws = wb['信源数据分析']
            
            # 找到品牌列的位置
            brand_columns = {}
            for row_idx in range(1, min(6, ws.max_row + 1)):
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value and isinstance(cell_value, str) and cell_value not in ['None', '关键词名称', 'AI平台', '信源平台名称']:
                        brand_name = cell_value
                        if brand_name not in brand_columns:
                            brand_columns[brand_name] = []
                        brand_columns[brand_name].append(col_idx)
            
            print(f"找到品牌: {list(brand_columns.keys())}")
            
            # 获取子标题（第2行）
            sub_headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=2, column=col_idx)
                sub_headers.append(cell.value)
            
            # 建立品牌到列的映射
            brand_to_columns = {}
            for brand_name, col_indices in brand_columns.items():
                if len(col_indices) > 1:  # 合并单元格
                    start_col = min(col_indices)
                    end_col = max(col_indices)
                    brand_sub_headers = sub_headers[start_col-1:end_col]
                else:  # 单列
                    start_col = col_indices[0]
                    end_col = col_indices[0]
                    brand_sub_headers = [sub_headers[start_col-1]] if start_col <= len(sub_headers) else []
                
                brand_to_columns[brand_name] = {
                    'start_col': start_col,
                    'end_col': end_col,
                    'sub_headers': brand_sub_headers
                }
            
            # 提取数据
            data_rows = []
            data_start_row = 3  # 数据从第3行开始
            
            for row_idx in range(data_start_row, ws.max_row + 1):
                # 获取关键词名称（第一列）
                keyword_name = ws.cell(row=row_idx, column=1).value
                ai_platform = ws.cell(row=row_idx, column=2).value
                source_platform = ws.cell(row=row_idx, column=3).value
                
                if keyword_name is None or keyword_name == '':
                    continue
                
                # 为每个品牌提取数据
                for brand_name, col_info in brand_to_columns.items():
                    row_data = {
                        '关键词名称': keyword_name,
                        'AI平台': ai_platform,
                        '信源平台名称': source_platform,
                        '品牌': brand_name
                    }
                    
                    # 提取该品牌对应的数据列
                    start_col = col_info['start_col']
                    end_col = col_info['end_col']
                    sub_headers = col_info['sub_headers']
                    
                    for i, sub_header in enumerate(sub_headers):
                        col_idx = start_col + i
                        if col_idx <= end_col and col_idx <= ws.max_column:
                            cell_value = ws.cell(row=row_idx, column=col_idx).value
                            row_data[sub_header] = cell_value
                        else:
                            row_data[sub_header] = None
                    
                    data_rows.append(row_data)
            
            df_sources = pd.DataFrame(data_rows)
            results['信源数据分析_转置'] = df_sources
            print(f"信源数据分析转置后形状: {df_sources.shape}")
        
        # 保存结果
        if output_file is None:
            base_name = os.path.splitext(input_file)[0]
            output_file = f"{base_name}_转置后.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df in results.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        file_size = os.path.getsize(output_file)
        print(f"\n文件已保存: {output_file}")
        print(f"文件大小: {file_size / 1024:.2f} KB")
        
        # 数据验证
        print("\n数据验证:")
        for sheet_name, df in results.items():
            print(f"\n{sheet_name}:")
            print(f"  行数: {len(df)}")
            print(f"  列数: {len(df.columns)}")
            if '品牌' in df.columns:
                print(f"  唯一品牌数: {df['品牌'].nunique()}")
            if '关键词名称' in df.columns:
                print(f"  唯一关键词数: {df['关键词名称'].nunique()}")
        
        return results
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """
    主函数 - 命令行使用
    """
    if len(sys.argv) < 2:
        print("使用方法: python process_three_test_report.py <输入文件> [输出文件]")
        print("示例: python process_three_test_report.py 三次测试对内报表.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    print("=" * 60)
    print("三次测试对内报表转置工具")
    print("=" * 60)
    
    results = process_three_test_report(input_file, output_file)
    
    if results is not None:
        print("\n" + "=" * 60)
        print("处理完成!")
        print("=" * 60)
    else:
        print("\n处理失败!")

if __name__ == "__main__":
    main()
