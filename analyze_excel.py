#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件结构分析工具
详细分析原始Excel文件的数据结构，找出正确的数据映射关系
"""

import pandas as pd
import openpyxl
import numpy as np

def analyze_excel_structure_detailed(file_path, sheet_name=None):
    """
    详细分析Excel文件结构
    """
    print(f"正在详细分析Excel文件: {file_path}")
    
    # 使用openpyxl读取工作簿
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 检查工作表
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    
    print(f"使用工作表: {sheet_name}")
    ws = wb[sheet_name]
    
    print(f"工作表尺寸: {ws.max_row} 行 x {ws.max_column} 列")
    
    # 分析前几行的结构
    print("\n=== 前5行数据分析 ===")
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(20, ws.max_column + 1)):  # 只看前20列
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(cell.value)
        print(f"第{row_idx}行: {row_data}")
    
    # 分析合并单元格
    merged_ranges = list(ws.merged_cells.ranges)
    print(f"\n=== 合并单元格分析 ===")
    print(f"找到 {len(merged_ranges)} 个合并单元格区域:")
    
    brand_columns = {}
    for merged_range in merged_ranges:
        top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
        brand_name = top_left_cell.value
        
        if brand_name and isinstance(brand_name, str):
            brand_columns[brand_name] = {
                'start_col': merged_range.min_col,
                'end_col': merged_range.max_col,
                'row': merged_range.min_row
            }
            print(f"品牌: {brand_name}")
            print(f"  列范围: {merged_range.min_col} - {merged_range.max_col}")
            print(f"  行: {merged_range.min_row}")
    
    # 分析子标题行（通常是第2行）
    print(f"\n=== 子标题行分析 ===")
    sub_headers_row = 2
    sub_headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=sub_headers_row, column=col_idx)
        sub_headers.append(cell.value)
    
    print(f"第{sub_headers_row}行子标题: {sub_headers[:20]}...")  # 只显示前20个
    
    # 分析数据行
    print(f"\n=== 数据行分析 ===")
    data_start_row = 3  # 假设数据从第3行开始
    
    # 查看前几行数据
    for row_idx in range(data_start_row, min(data_start_row + 3, ws.max_row + 1)):
        print(f"\n第{row_idx}行数据:")
        row_data = []
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(cell.value)
        print(f"  {row_data}")
    
    return wb, ws, brand_columns, sub_headers

def find_data_mapping(wb, ws, brand_columns, sub_headers):
    """
    找到正确的数据映射关系
    """
    print(f"\n=== 数据映射分析 ===")
    
    # 创建品牌到列索引的映射
    brand_to_columns = {}
    
    for brand_name, info in brand_columns.items():
        start_col = info['start_col']
        end_col = info['end_col']
        
        # 找到这个品牌对应的子列
        brand_sub_headers = []
        for i in range(start_col, end_col + 1):
            if i <= len(sub_headers):
                brand_sub_headers.append(sub_headers[i-1])
        
        brand_to_columns[brand_name] = {
            'start_col': start_col,
            'end_col': end_col,
            'sub_headers': brand_sub_headers
        }
        
        print(f"品牌: {brand_name}")
        print(f"  列范围: {start_col}-{end_col}")
        print(f"  子标题: {brand_sub_headers}")
    
    return brand_to_columns

def extract_data_correctly(ws, brand_to_columns):
    """
    正确提取数据
    """
    print(f"\n=== 数据提取 ===")
    
    data_rows = []
    data_start_row = 3  # 假设数据从第3行开始
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        # 获取信源平台名称（第一列）
        source_platform = ws.cell(row=row_idx, column=1).value
        
        if source_platform is None:
            continue
        
        print(f"处理信源平台: {source_platform}")
        
        # 为每个品牌提取数据
        for brand_name, col_info in brand_to_columns.items():
            row_data = {
                '信源平台名称': source_platform,
                '品牌': brand_name
            }
            
            # 提取该品牌对应的数据列
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            sub_headers = col_info['sub_headers']
            
            for i, sub_header in enumerate(sub_headers):
                col_idx = start_col + i
                if col_idx <= end_col and col_idx <= ws.max_column:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_data[sub_header] = cell_value
                    print(f"  {brand_name} - {sub_header}: {cell_value}")
                else:
                    row_data[sub_header] = None
            
            data_rows.append(row_data)
        
        # 只处理前几行作为示例
        if len(data_rows) >= 20:  # 限制输出数量
            break
    
    return data_rows

def main():
    """
    主函数
    """
    input_file = "2025916移山科技循环10次采集任务34词对外报表_待处理.xlsx"
    
    print("=" * 60)
    print("Excel文件结构详细分析工具")
    print("=" * 60)
    
    try:
        # 分析文件结构
        wb, ws, brand_columns, sub_headers = analyze_excel_structure_detailed(input_file)
        
        # 找到数据映射
        brand_to_columns = find_data_mapping(wb, ws, brand_columns, sub_headers)
        
        # 提取数据
        data_rows = extract_data_correctly(ws, brand_to_columns)
        
        # 显示提取的数据
        print(f"\n=== 提取的数据示例 ===")
        for i, row in enumerate(data_rows[:10]):  # 只显示前10行
            print(f"行{i+1}: {row}")
        
        print(f"\n总共提取了 {len(data_rows)} 行数据")
        
    except Exception as e:
        print(f"分析过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

