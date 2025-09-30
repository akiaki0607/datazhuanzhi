#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
思迈特报表转置处理脚本
处理2025918_927合并_思迈特对外_对内_报表_副本.xlsx文件
"""

import pandas as pd
import openpyxl
import os
from datetime import datetime

def transpose_ai_platform_sheet(ws):
    """转置AI平台的核心指标工作表"""
    print("处理AI平台的核心指标工作表...")
    
    # 识别合并单元格中的品牌
    brand_columns = {}
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
        brand_name = top_left_cell.value
        if brand_name and isinstance(brand_name, str) and brand_name.strip():
            brand_columns[brand_name] = {
                'start_col': merged_range.min_col,
                'end_col': merged_range.max_col
            }
    
    # 提取数据
    data_rows = []
    data_start_row = 3  # 数据从第3行开始
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        # 获取基础信息
        date = ws.cell(row=row_idx, column=1).value
        ai_platform = ws.cell(row=row_idx, column=2).value
        total_answers = ws.cell(row=row_idx, column=3).value
        
        if ai_platform is None or ai_platform == '':
            continue
        
        # 为每个品牌提取数据
        for brand_name, col_info in brand_columns.items():
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            
            # 检查该品牌是否有有效数据
            has_data = False
            for col_idx in range(start_col, end_col + 1):
                if col_idx <= ws.max_column:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and cell_value != '' and cell_value != 0:
                        has_data = True
                        break
            
            # 只有当品牌有数据时才添加行
            if has_data:
                row_data = {
                    '日期': date,
                    'AI平台名称': ai_platform,
                    'AI回答总条数': total_answers,
                    '品牌': brand_name.split('(')[0],  # 提取品牌名称
                    '品牌类型': '客户' if '客户' in brand_name else '竞品'
                }
                
                # 添加该品牌对应的数据列
                for col_idx in range(start_col, end_col + 1):
                    if col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if col_idx == start_col:
                            row_data['AI平台的可见占比'] = cell_value
                        elif col_idx == start_col + 1:
                            row_data['AI平台的推荐占比'] = cell_value
                
                data_rows.append(row_data)
    
    return pd.DataFrame(data_rows)

def transpose_keyword_sheet(ws):
    """转置关键词工作表"""
    print("处理关键词工作表...")
    
    # 识别合并单元格中的品牌
    brand_columns = {}
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
        brand_name = top_left_cell.value
        if brand_name and isinstance(brand_name, str) and brand_name.strip():
            brand_columns[brand_name] = {
                'start_col': merged_range.min_col,
                'end_col': merged_range.max_col
            }
    
    # 提取数据
    data_rows = []
    data_start_row = 3  # 数据从第3行开始
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        # 获取基础信息
        date = ws.cell(row=row_idx, column=1).value
        keyword = ws.cell(row=row_idx, column=2).value
        ai_platform = ws.cell(row=row_idx, column=3).value
        
        if keyword is None or keyword == '':
            continue
        
        # 为每个品牌提取数据
        for brand_name, col_info in brand_columns.items():
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            
            # 检查该品牌是否有有效数据
            has_data = False
            for col_idx in range(start_col, end_col + 1):
                if col_idx <= ws.max_column:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and cell_value != '' and cell_value != 0:
                        has_data = True
                        break
            
            # 只有当品牌有数据时才添加行
            if has_data:
                row_data = {
                    '日期': date,
                    '关键词名称': keyword,
                    'AI平台名称': ai_platform,
                    '品牌': brand_name.split('(')[0],  # 提取品牌名称
                    '品牌类型': '客户' if '客户' in brand_name else '竞品'
                }
                
                # 添加该品牌对应的数据列
                for col_idx in range(start_col, end_col + 1):
                    if col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if col_idx == start_col:
                            row_data['可见概率'] = cell_value
                        elif col_idx == start_col + 1:
                            row_data['推荐概率'] = cell_value
                        elif col_idx == start_col + 2:
                            row_data['信源平台占比'] = cell_value
                        elif col_idx == start_col + 3:
                            row_data['信源文章占比'] = cell_value
                        elif col_idx == start_col + 4:
                            row_data['Top1占比'] = cell_value
                        elif col_idx == start_col + 5:
                            row_data['Top前3占比'] = cell_value
                        elif col_idx == start_col + 6:
                            row_data['Top前5占比'] = cell_value
                        elif col_idx == start_col + 7:
                            row_data['Top前10占比'] = cell_value
                
                data_rows.append(row_data)
    
    return pd.DataFrame(data_rows)

def process_simait_excel_transpose(input_file_path, output_file_path):
    """处理思迈特Excel转置"""
    try:
        # 读取原始文件
        wb_original = openpyxl.load_workbook(input_file_path, data_only=True)
        print(f"原始文件工作表: {wb_original.sheetnames}")
        
        results = {}
        
        # 处理AI平台的核心指标工作表
        if "AI平台的核心指标" in wb_original.sheetnames:
            ws_original = wb_original["AI平台的核心指标"]
            df_transposed = transpose_ai_platform_sheet(ws_original)
            results["AI平台的核心指标"] = df_transposed
        
        # 处理关键词工作表
        if "关键词" in wb_original.sheetnames:
            ws_original = wb_original["关键词"]
            df_transposed = transpose_keyword_sheet(ws_original)
            results["关键词"] = df_transposed
        
        # 使用pandas保存所有工作表
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # 保存转置后的AI平台的核心指标
            if "AI平台的核心指标" in results and results["AI平台的核心指标"] is not None:
                results["AI平台的核心指标"].to_excel(writer, sheet_name='AI平台的核心指标', index=False)
                print(f"AI平台的核心指标转置完成: {results['AI平台的核心指标'].shape}")
            
            # 保存转置后的关键词
            if "关键词" in results and results["关键词"] is not None:
                results["关键词"].to_excel(writer, sheet_name='关键词', index=False)
                print(f"关键词转置完成: {results['关键词'].shape}")
            
            # 其他工作表直接复制
            for sheet_name in wb_original.sheetnames:
                if sheet_name not in ["AI平台的核心指标", "关键词"]:
                    ws_original = wb_original[sheet_name]
                    # 读取原工作表数据
                    data = []
                    for row in ws_original.iter_rows(values_only=True):
                        data.append(row)
                    
                    # 创建DataFrame并保存
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    print(f"复制工作表: {sheet_name}")
        
        return results
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    # 输入文件路径
    input_file = "/Users/aki/Documents/AI相关/cursor AI代码练习/转置-思迈特对外报表/待处理文件/2025918_927合并_思迈特对外_对内_报表_副本.xlsx"
    
    # 输出文件路径
    current_date = datetime.now().strftime("%Y%m%d")
    output_file = f"/Users/aki/Documents/AI相关/cursor AI代码练习/转置-思迈特对外报表/2025918_927合并_思迈特对外_对内_报表_副本_{current_date}_转置完成.xlsx"
    
    print(f"开始处理文件: {input_file}")
    print(f"输出文件: {output_file}")
    
    # 处理转置
    results = process_simait_excel_transpose(input_file, output_file)
    
    if results is not None:
        print("\n转置处理完成！")
        print("处理结果:")
        for sheet_name, df in results.items():
            if df is not None:
                print(f"  {sheet_name}: {df.shape[0]}行 × {df.shape[1]}列")
        print(f"\n输出文件已保存: {output_file}")
    else:
        print("转置处理失败！")
