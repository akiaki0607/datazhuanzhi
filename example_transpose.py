#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
示例文件转置处理工具
按照示例格式处理示例文件
"""

import pandas as pd
import openpyxl
import os
import sys
from datetime import datetime

def transpose_example_source_data(ws):
    """
    转置示例文件信源数据分析工作表
    按照示例格式：将品牌从列标题转换为行数据
    """
    print("处理示例文件信源数据分析工作表...")
    
    # 查找改动后数据
    after_data_start = None
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=2).value
        if cell_value == "改动后":
            # 查找表头行
            for header_row in range(row_idx + 1, ws.max_row + 1):
                header_cell = ws.cell(row=header_row, column=2).value
                if header_cell == "关键词名称":
                    after_data_start = header_row
                    break
            break
    
    if after_data_start is None:
        print("未找到改动后数据")
        return None
    
    print(f"改动后数据从第{after_data_start}行开始")
    
    # 提取改动后数据
    after_data = []
    for row_idx in range(after_data_start + 1, ws.max_row + 1):
        keyword = ws.cell(row=row_idx, column=2).value
        if keyword and keyword != '':
            ai_platform = ws.cell(row=row_idx, column=3).value
            source_platform = ws.cell(row=row_idx, column=4).value
            total_articles = ws.cell(row=row_idx, column=5).value
            brand = ws.cell(row=row_idx, column=6).value
            brand_type = ws.cell(row=row_idx, column=7).value
            ratio = ws.cell(row=row_idx, column=8).value
            count = ws.cell(row=row_idx, column=9).value
            
            row_data = {
                '关键词名称': keyword,
                'AI平台': ai_platform,
                '信源平台名称': source_platform,
                '选用信源文章总数': total_articles,
                '品牌': brand,
                '品牌类型': brand_type,
                '选用信源文章占比': ratio,
                '选用信源文章数': count
            }
            
            after_data.append(row_data)
    
    print(f"提取了 {len(after_data)} 行数据")
    return pd.DataFrame(after_data)

def process_example_transpose(input_file, output_file=None):
    """
    示例文件转置处理函数
    处理所有工作表，保持sheet数量一致
    
    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径（可选，默认自动生成）
    
    返回:
    dict: 各工作表的转换结果
    """
    try:
        print(f"开始处理文件: {input_file}")
        
        # 检查输入文件是否存在
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"找不到输入文件: {input_file}")
        
        # 读取原始文件
        wb_original = openpyxl.load_workbook(input_file, data_only=True)
        print(f"原始文件工作表: {wb_original.sheetnames}")
        
        # 创建输出文件
        if output_file is None:
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            current_date = datetime.now().strftime("%Y%m%d")
            output_file = f"{base_name}_{current_date}_示例转置完成.xlsx"
        
        results = {}
        
        # 处理信源数据分析工作表
        if "信源数据分析" in wb_original.sheetnames:
            ws_original = wb_original["信源数据分析"]
            df_transposed = transpose_example_source_data(ws_original)
            results["信源数据分析"] = df_transposed
        
        # 使用pandas保存所有工作表
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 保存转置后的信源数据分析
            if "信源数据分析" in results and results["信源数据分析"] is not None:
                results["信源数据分析"].to_excel(writer, sheet_name='信源数据分析', index=False)
                print(f"信源数据分析转置完成: {results['信源数据分析'].shape}")
            
            # 其他工作表直接复制
            for sheet_name in wb_original.sheetnames:
                if sheet_name != "信源数据分析":
                    ws_original = wb_original[sheet_name]
                    # 读取原工作表数据
                    data = []
                    for row in ws_original.iter_rows(values_only=True):
                        data.append(row)
                    
                    # 创建DataFrame并保存
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    print(f"复制工作表: {sheet_name}")
        
        # 保存文件
        file_size = os.path.getsize(output_file)
        print(f"\n文件已保存: {output_file}")
        print(f"文件大小: {file_size / 1024:.2f} KB")
        
        return results
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """
    主函数 - 命令行使用
    """
    if len(sys.argv) < 2:
        print("使用方法: python example_transpose.py <输入文件> [输出文件]")
        print("示例: python example_transpose.py 数据文件.xlsx")
        print("示例: python example_transpose.py 数据文件.xlsx 输出文件.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    print("=" * 60)
    print("示例文件转置处理工具")
    print("=" * 60)
    
    results = process_example_transpose(input_file, output_file)
    
    if results is not None:
        print("\n" + "=" * 60)
        print("处理完成!")
        print("=" * 60)
        for sheet_name, df in results.items():
            if df is not None:
                print(f"{sheet_name}: {df.shape}")
    else:
        print("\n处理失败!")

if __name__ == "__main__":
    main()

