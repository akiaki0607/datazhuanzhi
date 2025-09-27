#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
精确转置处理工具
按照示例格式精确处理所有工作表，保持sheet数量一致
"""

import pandas as pd
import openpyxl
import os
import sys
from datetime import datetime

def transpose_source_data_sheet(ws):
    """
    转置信源数据分析工作表
    按照示例格式：将品牌从列标题转换为行数据
    """
    print("处理信源数据分析工作表...")
    
    # 查找改动前数据
    before_data_start = None
    before_headers = None
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=2).value
        if cell_value == "改动前":
            # 查找表头行
            for header_row in range(row_idx + 1, ws.max_row + 1):
                header_cell = ws.cell(row=header_row, column=2).value
                if header_cell == "关键词名称":
                    before_data_start = header_row
                    # 获取表头
                    before_headers = []
                    for col_idx in range(2, ws.max_column + 1):
                        cell_value = ws.cell(row=header_row, column=col_idx).value
                        before_headers.append(cell_value)
                    break
            break
    
    if before_data_start is None:
        print("未找到改动前数据")
        return None
    
    print(f"改动前表头: {before_headers}")
    
    # 查找改动后数据
    after_data_start = None
    after_headers = None
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=2).value
        if cell_value == "改动后":
            # 查找表头行
            for header_row in range(row_idx + 1, ws.max_row + 1):
                header_cell = ws.cell(row=header_row, column=2).value
                if header_cell == "关键词名称":
                    after_data_start = header_row
                    # 获取表头
                    after_headers = []
                    for col_idx in range(2, ws.max_column + 1):
                        cell_value = ws.cell(row=header_row, column=col_idx).value
                        after_headers.append(cell_value)
                    break
            break
    
    if after_data_start is None:
        print("未找到改动后数据")
        return None
    
    print(f"改动后表头: {after_headers}")
    
    # 识别品牌列（从改动前数据中识别）
    brand_columns = {}
    for row_idx in range(1, before_data_start):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and isinstance(cell_value, str) and ("客户" in cell_value or "竞品" in cell_value):
                # 查找该品牌对应的列范围
                start_col = col_idx
                end_col = col_idx
                # 向后查找连续的空列
                for next_col in range(col_idx + 1, ws.max_column + 1):
                    next_cell = ws.cell(row=row_idx, column=next_col)
                    if next_cell.value is None or next_cell.value == '':
                        end_col = next_col
                    else:
                        break
                
                brand_columns[cell_value] = {
                    'start_col': start_col,
                    'end_col': end_col
                }
                print(f"  品牌: {cell_value} (列 {start_col}-{end_col})")
    
    # 提取改动前数据
    before_data = []
    for row_idx in range(before_data_start + 1, ws.max_row + 1):
        keyword = ws.cell(row=row_idx, column=2).value
        if keyword and keyword != "改动后":
            ai_platform = ws.cell(row=row_idx, column=3).value
            source_platform = ws.cell(row=row_idx, column=4).value
            total_articles = ws.cell(row=row_idx, column=5).value
            
            row_data = {
                '关键词名称': keyword,
                'AI平台': ai_platform,
                '信源平台名称': source_platform,
                '选用信源文章总数': total_articles
            }
            
            # 为每个品牌添加数据
            for brand_name, col_info in brand_columns.items():
                start_col = col_info['start_col']
                end_col = col_info['end_col']
                
                for col_idx in range(start_col, end_col + 1):
                    if col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if col_idx == start_col:
                            row_data[f'{brand_name}_选用信源文章占比'] = cell_value
                        elif col_idx == start_col + 1:
                            row_data[f'{brand_name}_选用信源文章数'] = cell_value
            
            before_data.append(row_data)
        elif keyword == "改动后":
            break
    
    print(f"改动前数据: {len(before_data)} 行")
    
    # 提取改动后数据
    after_data = []
    for row_idx in range(after_data_start + 1, ws.max_row + 1):
        keyword = ws.cell(row=row_idx, column=2).value
        if keyword:
            ai_platform = ws.cell(row=row_idx, column=3).value
            source_platform = ws.cell(row=row_idx, column=4).value
            total_articles = ws.cell(row=row_idx, column=5).value
            brand = ws.cell(row=row_idx, column=6).value
            brand_type = ws.cell(row=row_idx, column=7).value
            ratio = ws.cell(row=row_idx, column=8).value
            count = ws.cell(row=row_idx, column=9).value
            
            row_data = {
                '关键词名称': keyword,
                'AI平台': ai_platform,
                '信源平台名称': source_platform,
                '选用信源文章总数': total_articles,
                '品牌': brand,
                '品牌类型': brand_type,
                '选用信源文章占比': ratio,
                '选用信源文章数': count
            }
            
            after_data.append(row_data)
    
    print(f"改动后数据: {len(after_data)} 行")
    
    # 返回改动后的数据（转置结果）
    return pd.DataFrame(after_data)

def process_complete_transpose(input_file, output_file=None):
    """
    完整转置处理函数
    处理所有工作表，保持sheet数量一致
    
    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径（可选，默认自动生成）
    
    返回:
    dict: 各工作表的转换结果
    """
    try:
        print(f"开始处理文件: {input_file}")
        
        # 检查输入文件是否存在
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"找不到输入文件: {input_file}")
        
        # 读取原始文件
        wb_original = openpyxl.load_workbook(input_file, data_only=True)
        print(f"原始文件工作表: {wb_original.sheetnames}")
        
        # 创建输出文件
        if output_file is None:
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            current_date = datetime.now().strftime("%Y%m%d")
            output_file = f"{base_name}_{current_date}_精确转置完成.xlsx"
        
        results = {}
        
        # 处理信源数据分析工作表
        if "信源数据分析" in wb_original.sheetnames:
            ws_original = wb_original["信源数据分析"]
            df_transposed = transpose_source_data_sheet(ws_original)
            results["信源数据分析"] = df_transposed
        
        # 使用pandas保存所有工作表
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 保存转置后的信源数据分析
            if "信源数据分析" in results and results["信源数据分析"] is not None:
                results["信源数据分析"].to_excel(writer, sheet_name='信源数据分析', index=False)
                print(f"信源数据分析转置完成: {results['信源数据分析'].shape}")
            
            # 其他工作表直接复制
            for sheet_name in wb_original.sheetnames:
                if sheet_name != "信源数据分析":
                    ws_original = wb_original[sheet_name]
                    # 读取原工作表数据
                    data = []
                    for row in ws_original.iter_rows(values_only=True):
                        data.append(row)
                    
                    # 创建DataFrame并保存
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    print(f"复制工作表: {sheet_name}")
        
        # 保存文件
        file_size = os.path.getsize(output_file)
        print(f"\n文件已保存: {output_file}")
        print(f"文件大小: {file_size / 1024:.2f} KB")
        
        return results
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """
    主函数 - 命令行使用
    """
    if len(sys.argv) < 2:
        print("使用方法: python precise_transpose.py <输入文件> [输出文件]")
        print("示例: python precise_transpose.py 数据文件.xlsx")
        print("示例: python precise_transpose.py 数据文件.xlsx 输出文件.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    print("=" * 60)
    print("精确转置处理工具")
    print("=" * 60)
    
    results = process_complete_transpose(input_file, output_file)
    
    if results is not None:
        print("\n" + "=" * 60)
        print("处理完成!")
        print("=" * 60)
        for sheet_name, df in results.items():
            if df is not None:
                print(f"{sheet_name}: {df.shape}")
    else:
        print("\n处理失败!")

if __name__ == "__main__":
    main()

