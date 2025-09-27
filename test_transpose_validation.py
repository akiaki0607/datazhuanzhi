#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
转置数据验证测试工具
自动测试转置后的数据是否正确
"""

import pandas as pd
import openpyxl
import os
import sys
from datetime import datetime

class TransposeValidator:
    """转置数据验证器"""
    
    def __init__(self, original_file, transposed_file):
        """
        初始化验证器
        
        参数:
        original_file: 原始Excel文件路径
        transposed_file: 转置后Excel文件路径
        """
        self.original_file = original_file
        self.transposed_file = transposed_file
        self.test_results = []
        
    def log_test(self, test_name, passed, message="", details=None):
        """记录测试结果"""
        result = {
            'test_name': test_name,
            'passed': passed,
            'message': message,
            'details': details,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        self.test_results.append(result)
        
        status = "✅ 通过" if passed else "❌ 失败"
        print(f"{status} {test_name}: {message}")
        if details:
            print(f"    详情: {details}")
    
    def test_file_existence(self):
        """测试文件是否存在"""
        test_name = "文件存在性检查"
        
        original_exists = os.path.exists(self.original_file)
        transposed_exists = os.path.exists(self.transposed_file)
        
        if original_exists and transposed_exists:
            self.log_test(test_name, True, "原始文件和转置文件都存在")
        else:
            missing_files = []
            if not original_exists:
                missing_files.append(f"原始文件: {self.original_file}")
            if not transposed_exists:
                missing_files.append(f"转置文件: {self.transposed_file}")
            
            self.log_test(test_name, False, f"文件缺失: {', '.join(missing_files)}")
    
    def test_data_structure(self):
        """测试数据结构"""
        test_name = "数据结构验证"
        
        try:
            # 读取转置后的数据
            df_transposed = pd.read_excel(self.transposed_file, sheet_name='转置后数据')
            
            # 检查必要的列是否存在
            required_columns = ['关键词名称', 'AI平台名称', '信源平台名称', '品牌']
            missing_columns = [col for col in required_columns if col not in df_transposed.columns]
            
            if not missing_columns:
                self.log_test(test_name, True, f"数据结构正确，包含{len(df_transposed.columns)}列")
                return df_transposed
            else:
                self.log_test(test_name, False, f"缺少必要列: {missing_columns}")
                return None
                
        except Exception as e:
            self.log_test(test_name, False, f"读取转置文件失败: {str(e)}")
            return None
    
    def test_data_completeness(self, df_transposed):
        """测试数据完整性"""
        test_name = "数据完整性验证"
        
        if df_transposed is None:
            self.log_test(test_name, False, "无法进行完整性验证，数据结构测试失败")
            return
        
        try:
            # 检查关键字段的非空数据
            keyword_count = df_transposed['关键词名称'].notna().sum()
            ai_platform_count = df_transposed['AI平台名称'].notna().sum()
            source_platform_count = df_transposed['信源平台名称'].notna().sum()
            brand_count = df_transposed['品牌'].notna().sum()
            
            total_rows = len(df_transposed)
            
            # 检查数据完整性
            completeness_checks = [
                (keyword_count > 0, f"关键词数据: {keyword_count}/{total_rows}"),
                (ai_platform_count > 0, f"AI平台数据: {ai_platform_count}/{total_rows}"),
                (source_platform_count > 0, f"信源平台数据: {source_platform_count}/{total_rows}"),
                (brand_count > 0, f"品牌数据: {brand_count}/{total_rows}")
            ]
            
            all_passed = all(check[0] for check in completeness_checks)
            details = "; ".join(check[1] for check in completeness_checks)
            
            if all_passed:
                self.log_test(test_name, True, f"数据完整性良好，总行数: {total_rows}", details)
            else:
                self.log_test(test_name, False, "数据完整性存在问题", details)
                
        except Exception as e:
            self.log_test(test_name, False, f"完整性验证失败: {str(e)}")
    
    def test_data_consistency(self, df_transposed):
        """测试数据一致性"""
        test_name = "数据一致性验证"
        
        if df_transposed is None:
            self.log_test(test_name, False, "无法进行一致性验证，数据结构测试失败")
            return
        
        try:
            # 检查唯一值数量
            unique_keywords = df_transposed['关键词名称'].nunique()
            unique_ai_platforms = df_transposed['AI平台名称'].nunique()
            unique_source_platforms = df_transposed['信源平台名称'].nunique()
            unique_brands = df_transposed['品牌'].nunique()
            
            # 检查数据范围是否合理
            consistency_checks = [
                (unique_keywords > 0, f"唯一关键词: {unique_keywords}"),
                (unique_ai_platforms > 0, f"唯一AI平台: {unique_ai_platforms}"),
                (unique_source_platforms > 0, f"唯一信源平台: {unique_source_platforms}"),
                (unique_brands > 0, f"唯一品牌: {unique_brands}")
            ]
            
            all_passed = all(check[0] for check in consistency_checks)
            details = "; ".join(check[1] for check in consistency_checks)
            
            if all_passed:
                self.log_test(test_name, True, "数据一致性良好", details)
            else:
                self.log_test(test_name, False, "数据一致性问题", details)
                
        except Exception as e:
            self.log_test(test_name, False, f"一致性验证失败: {str(e)}")
    
    def test_original_vs_transposed(self, df_transposed):
        """测试原始数据与转置数据的对应关系"""
        test_name = "原始数据与转置数据对应关系"
        
        if df_transposed is None:
            self.log_test(test_name, False, "无法进行对应关系验证，数据结构测试失败")
            return
        
        try:
            # 读取原始文件
            wb_original = openpyxl.load_workbook(self.original_file, data_only=True)
            
            # 找到信源数据分析工作表
            if "信源数据分析" in wb_original.sheetnames:
                ws_original = wb_original["信源数据分析"]
                
                # 计算原始数据行数（排除表头）
                original_data_rows = ws_original.max_row - 2  # 减去表头2行
                transposed_rows = len(df_transposed)
                
                # 计算期望的转置行数
                # 原始数据行数 × 品牌数量
                unique_brands = df_transposed['品牌'].nunique()
                expected_rows = original_data_rows * unique_brands
                
                # 允许一定的误差范围（±10%）
                tolerance = 0.1
                min_expected = int(expected_rows * (1 - tolerance))
                max_expected = int(expected_rows * (1 + tolerance))
                
                if min_expected <= transposed_rows <= max_expected:
                    self.log_test(test_name, True, 
                                f"转置行数合理: {transposed_rows} (期望: {expected_rows}±{tolerance*100:.0f}%)",
                                f"原始数据行: {original_data_rows}, 品牌数: {unique_brands}")
                else:
                    self.log_test(test_name, False, 
                                f"转置行数异常: {transposed_rows} (期望: {expected_rows}±{tolerance*100:.0f}%)",
                                f"原始数据行: {original_data_rows}, 品牌数: {unique_brands}")
            else:
                self.log_test(test_name, False, "原始文件中未找到'信源数据分析'工作表")
                
        except Exception as e:
            self.log_test(test_name, False, f"对应关系验证失败: {str(e)}")
    
    def test_data_quality(self, df_transposed):
        """测试数据质量"""
        test_name = "数据质量验证"
        
        if df_transposed is None:
            self.log_test(test_name, False, "无法进行质量验证，数据结构测试失败")
            return
        
        try:
            # 检查重复数据
            duplicate_rows = df_transposed.duplicated().sum()
            
            # 检查空值
            null_counts = df_transposed.isnull().sum()
            total_nulls = null_counts.sum()
            
            # 检查数据质量指标
            quality_checks = [
                (duplicate_rows == 0, f"重复行: {duplicate_rows}"),
                (total_nulls < len(df_transposed) * 0.5, f"空值总数: {total_nulls}"),
                (len(df_transposed) > 0, f"总行数: {len(df_transposed)}")
            ]
            
            all_passed = all(check[0] for check in quality_checks)
            details = "; ".join(check[1] for check in quality_checks)
            
            if all_passed:
                self.log_test(test_name, True, "数据质量良好", details)
            else:
                self.log_test(test_name, False, "数据质量问题", details)
                
        except Exception as e:
            self.log_test(test_name, False, f"质量验证失败: {str(e)}")
    
    def run_all_tests(self):
        """运行所有测试"""
        print("=" * 60)
        print("转置数据验证测试开始")
        print("=" * 60)
        
        # 测试1: 文件存在性
        self.test_file_existence()
        
        # 测试2: 数据结构
        df_transposed = self.test_data_structure()
        
        # 测试3: 数据完整性
        self.test_data_completeness(df_transposed)
        
        # 测试4: 数据一致性
        self.test_data_consistency(df_transposed)
        
        # 测试5: 原始数据与转置数据对应关系
        self.test_original_vs_transposed(df_transposed)
        
        # 测试6: 数据质量
        self.test_data_quality(df_transposed)
        
        # 生成测试报告
        return self.generate_report()
    
    def generate_report(self):
        """生成测试报告"""
        print("\n" + "=" * 60)
        print("测试报告")
        print("=" * 60)
        
        total_tests = len(self.test_results)
        passed_tests = sum(1 for result in self.test_results if result['passed'])
        failed_tests = total_tests - passed_tests
        
        print(f"总测试数: {total_tests}")
        print(f"通过测试: {passed_tests}")
        print(f"失败测试: {failed_tests}")
        print(f"通过率: {passed_tests/total_tests*100:.1f}%")
        
        if failed_tests > 0:
            print("\n失败的测试:")
            for result in self.test_results:
                if not result['passed']:
                    print(f"  ❌ {result['test_name']}: {result['message']}")
        
        # 保存测试报告到文件
        report_file = f"转置验证测试报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write("转置数据验证测试报告\n")
            f.write("=" * 50 + "\n")
            f.write(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"原始文件: {self.original_file}\n")
            f.write(f"转置文件: {self.transposed_file}\n")
            f.write(f"总测试数: {total_tests}\n")
            f.write(f"通过测试: {passed_tests}\n")
            f.write(f"失败测试: {failed_tests}\n")
            f.write(f"通过率: {passed_tests/total_tests*100:.1f}%\n\n")
            
            f.write("详细测试结果:\n")
            f.write("-" * 30 + "\n")
            for result in self.test_results:
                status = "通过" if result['passed'] else "失败"
                f.write(f"{status} {result['test_name']}: {result['message']}\n")
                if result['details']:
                    f.write(f"  详情: {result['details']}\n")
                f.write(f"  时间: {result['timestamp']}\n\n")
        
        print(f"\n测试报告已保存到: {report_file}")
        
        return passed_tests == total_tests

def main():
    """主函数"""
    if len(sys.argv) < 3:
        print("使用方法: python test_transpose_validation.py <原始文件> <转置文件>")
        print("示例: python test_transpose_validation.py 原始文件.xlsx 转置文件.xlsx")
        return
    
    original_file = sys.argv[1]
    transposed_file = sys.argv[2]
    
    validator = TransposeValidator(original_file, transposed_file)
    success = validator.run_all_tests()
    
    if success:
        print("\n🎉 所有测试通过！转置数据验证成功！")
        sys.exit(0)
    else:
        print("\n⚠️  部分测试失败，请检查转置结果！")
        sys.exit(1)

if __name__ == "__main__":
    main()
