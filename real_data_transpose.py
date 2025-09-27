#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
实际数据转置处理工具
按照示例格式处理实际的待处理数据
"""

import pandas as pd
import openpyxl
import os
import sys
from datetime import datetime

def transpose_source_data_sheet(ws):
    """
    转置信源数据分析工作表
    按照示例格式：将品牌从列标题转换为行数据
    """
    print("处理信源数据分析工作表...")
    
    # 识别合并单元格中的品牌
    brand_columns = {}
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
        brand_name = top_left_cell.value
        if brand_name and isinstance(brand_name, str) and brand_name.strip():
            brand_columns[brand_name] = {
                'start_col': merged_range.min_col,
                'end_col': merged_range.max_col
            }
            print(f"  品牌: {brand_name} (列 {merged_range.min_col}-{merged_range.max_col})")
    
    # 获取表头
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_idx)
        headers.append(cell.value)
    
    print(f"表头: {headers[:10]}...")  # 显示前10个
    
    # 提取数据
    data_rows = []
    data_start_row = 3  # 数据从第3行开始
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        # 获取基础信息
        keyword = ws.cell(row=row_idx, column=1).value
        ai_platform = ws.cell(row=row_idx, column=2).value
        source_platform = ws.cell(row=row_idx, column=3).value
        total_articles = ws.cell(row=row_idx, column=4).value
        
        if keyword is None or keyword == '':
            continue
        
        # 为每个品牌提取数据
        for brand_name, col_info in brand_columns.items():
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            
            # 检查该品牌是否有有效数据
            has_data = False
            for col_idx in range(start_col, end_col + 1):
                if col_idx <= ws.max_column:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and cell_value != '' and cell_value != 0:
                        has_data = True
                        break
            
            # 只有当品牌有数据时才添加行
            if has_data:
                row_data = {
                    '关键词名称': keyword,
                    'AI平台': ai_platform,
                    '信源平台名称': source_platform,
                    '选用信源文章总数': total_articles,
                    '品牌': brand_name.split('(')[0],  # 提取品牌名称
                    '品牌类型': '客户' if '客户' in brand_name else '竞品'
                }
                
                # 添加该品牌对应的数据列
                for col_idx in range(start_col, end_col + 1):
                    if col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if col_idx == start_col:
                            row_data['选用信源文章占比'] = cell_value
                        elif col_idx == start_col + 1:
                            row_data['选用信源文章数'] = cell_value
                
                data_rows.append(row_data)
        
        # 显示进度
        if row_idx % 100 == 0:
            print(f"已处理 {row_idx - data_start_row + 1} 行数据...")
    
    print(f"数据提取完成，总共 {len(data_rows)} 行数据")
    return pd.DataFrame(data_rows)

def process_real_data_transpose(input_file, output_file=None):
    """
    实际数据转置处理函数
    处理所有工作表，保持sheet数量一致
    
    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径（可选，默认自动生成）
    
    返回:
    dict: 各工作表的转换结果
    """
    try:
        print(f"开始处理文件: {input_file}")
        
        # 检查输入文件是否存在
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"找不到输入文件: {input_file}")
        
        # 读取原始文件
        wb_original = openpyxl.load_workbook(input_file, data_only=True)
        print(f"原始文件工作表: {wb_original.sheetnames}")
        
        # 创建输出文件
        if output_file is None:
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            current_date = datetime.now().strftime("%Y%m%d")
            output_file = f"{base_name}_{current_date}_实际转置完成.xlsx"
        
        results = {}
        
        # 处理信源数据分析工作表
        if "信源数据分析" in wb_original.sheetnames:
            ws_original = wb_original["信源数据分析"]
            df_transposed = transpose_source_data_sheet(ws_original)
            results["信源数据分析"] = df_transposed
        
        # 使用pandas保存所有工作表
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 保存转置后的信源数据分析
            if "信源数据分析" in results and results["信源数据分析"] is not None:
                results["信源数据分析"].to_excel(writer, sheet_name='信源数据分析', index=False)
                print(f"信源数据分析转置完成: {results['信源数据分析'].shape}")
            
            # 其他工作表直接复制
            for sheet_name in wb_original.sheetnames:
                if sheet_name != "信源数据分析":
                    ws_original = wb_original[sheet_name]
                    # 读取原工作表数据
                    data = []
                    for row in ws_original.iter_rows(values_only=True):
                        data.append(row)
                    
                    # 创建DataFrame并保存
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    print(f"复制工作表: {sheet_name}")
        
        # 保存文件
        file_size = os.path.getsize(output_file)
        print(f"\n文件已保存: {output_file}")
        print(f"文件大小: {file_size / 1024:.2f} KB")
        
        return results
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """
    主函数 - 命令行使用
    """
    if len(sys.argv) < 2:
        print("使用方法: python real_data_transpose.py <输入文件> [输出文件]")
        print("示例: python real_data_transpose.py 数据文件.xlsx")
        print("示例: python real_data_transpose.py 数据文件.xlsx 输出文件.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    print("=" * 60)
    print("实际数据转置处理工具")
    print("=" * 60)
    
    results = process_real_data_transpose(input_file, output_file)
    
    if results is not None:
        print("\n" + "=" * 60)
        print("处理完成!")
        print("=" * 60)
        for sheet_name, df in results.items():
            if df is not None:
                print(f"{sheet_name}: {df.shape}")
    else:
        print("\n处理失败!")

if __name__ == "__main__":
    main()

