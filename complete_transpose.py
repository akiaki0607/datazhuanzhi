#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完整转置处理工具
按照示例格式处理所有工作表，保持sheet数量一致
"""

import pandas as pd
import openpyxl
import os
import sys
from datetime import datetime

def transpose_source_data_sheet(ws):
    """
    转置信源数据分析工作表
    """
    print("处理信源数据分析工作表...")
    
    # 查找数据开始行
    data_start_row = None
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=2).value
        if cell_value == "关键词名称":
            data_start_row = row_idx
            break
    
    if data_start_row is None:
        print("未找到数据开始行")
        return None
    
    # 获取表头
    headers = []
    for col_idx in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=data_start_row, column=col_idx).value
        headers.append(cell_value)
    
    print(f"表头: {headers}")
    
    # 识别品牌列
    brand_columns = {}
    for row_idx in range(1, data_start_row):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and isinstance(cell_value, str) and ("客户" in cell_value or "竞品" in cell_value):
                # 查找该品牌对应的列范围
                start_col = col_idx
                end_col = col_idx
                # 向后查找连续的空列
                for next_col in range(col_idx + 1, ws.max_column + 1):
                    next_cell = ws.cell(row=row_idx, column=next_col)
                    if next_cell.value is None or next_cell.value == '':
                        end_col = next_col
                    else:
                        break
                
                brand_columns[cell_value] = {
                    'start_col': start_col,
                    'end_col': end_col
                }
                print(f"  品牌: {cell_value} (列 {start_col}-{end_col})")
    
    # 提取数据
    data_rows = []
    for row_idx in range(data_start_row + 1, ws.max_row + 1):
        # 获取基础信息
        keyword = ws.cell(row=row_idx, column=2).value
        ai_platform = ws.cell(row=row_idx, column=3).value
        source_platform = ws.cell(row=row_idx, column=4).value
        total_articles = ws.cell(row=row_idx, column=5).value
        
        if keyword is None or keyword == '':
            continue
        
        # 为每个品牌提取数据
        for brand_name, col_info in brand_columns.items():
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            
            # 检查该品牌是否有有效数据
            has_data = False
            for col_idx in range(start_col, end_col + 1):
                if col_idx <= ws.max_column:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and cell_value != '' and cell_value != 0:
                        has_data = True
                        break
            
            # 只有当品牌有数据时才添加行
            if has_data:
                row_data = {
                    '关键词名称': keyword,
                    'AI平台': ai_platform,
                    '信源平台名称': source_platform,
                    '选用信源文章总数': total_articles,
                    '品牌': brand_name.split('(')[0],  # 提取品牌名称
                    '品牌类型': '客户' if '客户' in brand_name else '竞品'
                }
                
                # 添加该品牌对应的数据列
                for col_idx in range(start_col, end_col + 1):
                    if col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if col_idx == start_col:
                            row_data['选用信源文章占比'] = cell_value
                        elif col_idx == start_col + 1:
                            row_data['选用信源文章数'] = cell_value
                
                data_rows.append(row_data)
    
    print(f"提取了 {len(data_rows)} 行数据")
    return pd.DataFrame(data_rows)

def process_complete_transpose(input_file, output_file=None):
    """
    完整转置处理函数
    处理所有工作表，保持sheet数量一致
    
    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径（可选，默认自动生成）
    
    返回:
    dict: 各工作表的转换结果
    """
    try:
        print(f"开始处理文件: {input_file}")
        
        # 检查输入文件是否存在
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"找不到输入文件: {input_file}")
        
        # 读取原始文件
        wb_original = openpyxl.load_workbook(input_file, data_only=True)
        print(f"原始文件工作表: {wb_original.sheetnames}")
        
        # 创建输出文件
        if output_file is None:
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            current_date = datetime.now().strftime("%Y%m%d")
            output_file = f"{base_name}_{current_date}_完整转置完成.xlsx"
        
        # 创建新的工作簿
        wb_output = openpyxl.Workbook()
        # 删除默认工作表
        wb_output.remove(wb_output.active)
        
        results = {}
        
        # 处理每个工作表
        for sheet_name in wb_original.sheetnames:
            print(f"\n处理工作表: {sheet_name}")
            ws_original = wb_original[sheet_name]
            
            if sheet_name == "信源数据分析":
                # 转置信源数据分析工作表
                df_transposed = transpose_source_data_sheet(ws_original)
                if df_transposed is not None:
                    # 创建新工作表
                    ws_output = wb_output.create_sheet(sheet_name)
                    
                    # 写入表头
                    headers = list(df_transposed.columns)
                    for col_idx, header in enumerate(headers, 1):
                        ws_output.cell(row=1, column=col_idx, value=header)
                    
                    # 写入数据
                    for row_idx, row_data in enumerate(df_transposed.itertuples(index=False), 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws_output.cell(row=row_idx, column=col_idx, value=value)
                    
                    results[sheet_name] = df_transposed
                    print(f"转置完成: {df_transposed.shape}")
                else:
                    # 如果转置失败，复制原工作表
                    ws_output = wb_output.create_sheet(sheet_name)
                    for row in ws_original.iter_rows():
                        for cell in row:
                            ws_output.cell(row=cell.row, column=cell.column, value=cell.value)
                    print("转置失败，复制原工作表")
            else:
                # 其他工作表直接复制
                ws_output = wb_output.create_sheet(sheet_name)
                for row in ws_original.iter_rows():
                    for cell in row:
                        ws_output.cell(row=cell.row, column=cell.column, value=cell.value)
                print("直接复制原工作表")
        
        # 保存文件
        wb_output.save(output_file)
        file_size = os.path.getsize(output_file)
        print(f"\n文件已保存: {output_file}")
        print(f"文件大小: {file_size / 1024:.2f} KB")
        
        return results
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """
    主函数 - 命令行使用
    """
    if len(sys.argv) < 2:
        print("使用方法: python complete_transpose.py <输入文件> [输出文件]")
        print("示例: python complete_transpose.py 数据文件.xlsx")
        print("示例: python complete_transpose.py 数据文件.xlsx 输出文件.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    print("=" * 60)
    print("完整转置处理工具")
    print("=" * 60)
    
    results = process_complete_transpose(input_file, output_file)
    
    if results is not None:
        print("\n" + "=" * 60)
        print("处理完成!")
        print("=" * 60)
        for sheet_name, df in results.items():
            if df is not None:
                print(f"{sheet_name}: {df.shape}")
    else:
        print("\n处理失败!")

if __name__ == "__main__":
    main()

