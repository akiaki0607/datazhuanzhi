#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel转置处理Web应用
支持上传Excel文件，按照要求转置表格，处理完后可以下载
"""

from flask import Flask, request, render_template, send_file, jsonify, flash, redirect, url_for
import pandas as pd
import openpyxl
import os
import tempfile
import uuid
from datetime import datetime
from werkzeug.utils import secure_filename
import traceback

app = Flask(__name__)
app.secret_key = 'excel_transpose_secret_key_2025'

# 配置
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# 确保目录存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def transpose_source_data_sheet(ws):
    """转置信源数据分析工作表"""
    print("处理信源数据分析工作表...")
    
    # 识别合并单元格中的品牌
    brand_columns = {}
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
        brand_name = top_left_cell.value
        if brand_name and isinstance(brand_name, str) and brand_name.strip():
            brand_columns[brand_name] = {
                'start_col': merged_range.min_col,
                'end_col': merged_range.max_col
            }
    
    # 提取数据
    data_rows = []
    data_start_row = 3  # 数据从第3行开始
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        # 获取基础信息
        keyword = ws.cell(row=row_idx, column=1).value
        ai_platform = ws.cell(row=row_idx, column=2).value
        source_platform = ws.cell(row=row_idx, column=3).value
        total_articles = ws.cell(row=row_idx, column=4).value
        
        if keyword is None or keyword == '':
            continue
        
        # 为每个品牌提取数据
        for brand_name, col_info in brand_columns.items():
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            
            # 检查该品牌是否有有效数据
            has_data = False
            for col_idx in range(start_col, end_col + 1):
                if col_idx <= ws.max_column:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and cell_value != '' and cell_value != 0:
                        has_data = True
                        break
            
            # 只有当品牌有数据时才添加行
            if has_data:
                row_data = {
                    '关键词名称': keyword,
                    'AI平台': ai_platform,
                    '信源平台名称': source_platform,
                    '选用信源文章总数': total_articles,
                    '品牌': brand_name.split('(')[0],  # 提取品牌名称
                    '品牌类型': '客户' if '客户' in brand_name else '竞品'
                }
                
                # 添加该品牌对应的数据列
                for col_idx in range(start_col, end_col + 1):
                    if col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if col_idx == start_col:
                            row_data['选用信源文章占比'] = cell_value
                        elif col_idx == start_col + 1:
                            row_data['选用信源文章数'] = cell_value
                
                data_rows.append(row_data)
    
    return pd.DataFrame(data_rows)

def transpose_keyword_data_sheet(ws):
    """转置关键词数据分析工作表"""
    print("处理关键词数据分析工作表...")
    
    # 识别合并单元格中的品牌
    brand_columns = {}
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
        brand_name = top_left_cell.value
        if brand_name and isinstance(brand_name, str) and brand_name.strip():
            brand_columns[brand_name] = {
                'start_col': merged_range.min_col,
                'end_col': merged_range.max_col
            }
    
    # 提取数据
    data_rows = []
    data_start_row = 3  # 数据从第3行开始
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        # 获取基础信息
        keyword = ws.cell(row=row_idx, column=1).value
        ai_platform = ws.cell(row=row_idx, column=2).value
        
        if keyword is None or keyword == '':
            continue
        
        # 为每个品牌提取数据
        for brand_name, col_info in brand_columns.items():
            start_col = col_info['start_col']
            end_col = col_info['end_col']
            
            # 检查该品牌是否有有效数据
            has_data = False
            for col_idx in range(start_col, end_col + 1):
                if col_idx <= ws.max_column:
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and cell_value != '' and cell_value != 0:
                        has_data = True
                        break
            
            # 只有当品牌有数据时才添加行
            if has_data:
                row_data = {
                    '关键词名称': keyword,
                    'AI平台名称': ai_platform,
                    '品牌': brand_name.split('(')[0],  # 提取品牌名称
                    '品牌类型': '客户' if '客户' in brand_name else '竞品'
                }
                
                # 添加该品牌对应的数据列
                for col_idx in range(start_col, end_col + 1):
                    if col_idx <= ws.max_column:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if col_idx == start_col:
                            row_data['可见概率'] = cell_value
                        elif col_idx == start_col + 1:
                            row_data['推荐概率'] = cell_value
                        elif col_idx == start_col + 2:
                            row_data['信源平台占比'] = cell_value
                        elif col_idx == start_col + 3:
                            row_data['信源文章占比'] = cell_value
                        elif col_idx == start_col + 4:
                            row_data['Top1占比'] = cell_value
                        elif col_idx == start_col + 5:
                            row_data['Top前3占比'] = cell_value
                        elif col_idx == start_col + 6:
                            row_data['Top前5占比'] = cell_value
                        elif col_idx == start_col + 7:
                            row_data['Top前10占比'] = cell_value
                
                data_rows.append(row_data)
    
    return pd.DataFrame(data_rows)

def process_excel_transpose(input_file_path, output_file_path):
    """处理Excel转置"""
    try:
        # 读取原始文件
        wb_original = openpyxl.load_workbook(input_file_path, data_only=True)
        print(f"原始文件工作表: {wb_original.sheetnames}")
        
        results = {}
        
        # 处理信源数据分析工作表
        if "信源数据分析" in wb_original.sheetnames:
            ws_original = wb_original["信源数据分析"]
            df_transposed = transpose_source_data_sheet(ws_original)
            results["信源数据分析"] = df_transposed
        
        # 处理关键词数据分析工作表
        if "关键词数据分析" in wb_original.sheetnames:
            ws_original = wb_original["关键词数据分析"]
            df_transposed = transpose_keyword_data_sheet(ws_original)
            results["关键词数据分析"] = df_transposed
        
        # 使用pandas保存所有工作表
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # 保存转置后的信源数据分析
            if "信源数据分析" in results and results["信源数据分析"] is not None:
                results["信源数据分析"].to_excel(writer, sheet_name='信源数据分析', index=False)
                print(f"信源数据分析转置完成: {results['信源数据分析'].shape}")
            
            # 保存转置后的关键词数据分析
            if "关键词数据分析" in results and results["关键词数据分析"] is not None:
                results["关键词数据分析"].to_excel(writer, sheet_name='关键词数据分析', index=False)
                print(f"关键词数据分析转置完成: {results['关键词数据分析'].shape}")
            
            # 其他工作表直接复制
            for sheet_name in wb_original.sheetnames:
                if sheet_name not in ["信源数据分析", "关键词数据分析"]:
                    ws_original = wb_original[sheet_name]
                    # 读取原工作表数据
                    data = []
                    for row in ws_original.iter_rows(values_only=True):
                        data.append(row)
                    
                    # 创建DataFrame并保存
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    print(f"复制工作表: {sheet_name}")
        
        return results
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        traceback.print_exc()
        return None

@app.route('/')
def index():
    """主页"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """上传文件处理"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有选择文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if file and allowed_file(file.filename):
            # 生成唯一文件名
            filename = secure_filename(file.filename)
            unique_id = str(uuid.uuid4())
            input_filename = f"{unique_id}_{filename}"
            input_path = os.path.join(UPLOAD_FOLDER, input_filename)
            
            # 保存上传的文件
            file.save(input_path)
            
            # 生成输出文件名
            base_name = os.path.splitext(filename)[0]
            current_date = datetime.now().strftime("%Y%m%d")
            output_filename = f"{base_name}_{current_date}_完整转置完成.xlsx"
            output_path = os.path.join(OUTPUT_FOLDER, f"{unique_id}_{output_filename}")
            
            # 处理转置
            results = process_excel_transpose(input_path, output_path)
            
            if results is not None:
                # 清理上传文件
                os.remove(input_path)
                
                return jsonify({
                    'success': True,
                    'message': '转置处理完成',
                    'download_url': f'/download/{unique_id}_{output_filename}',
                    'results': {
                        sheet_name: df.shape for sheet_name, df in results.items() if df is not None
                    }
                })
            else:
                # 清理上传文件
                os.remove(input_path)
                return jsonify({'error': '转置处理失败'}), 500
        else:
            return jsonify({'error': '不支持的文件格式，请上传.xlsx或.xls文件'}), 400
            
    except Exception as e:
        return jsonify({'error': f'处理过程中出现错误: {str(e)}'}), 500

@app.route('/download/<filename>')
def download_file(filename):
    """下载文件"""
    try:
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=filename)
        else:
            return jsonify({'error': '文件不存在'}), 404
    except Exception as e:
        return jsonify({'error': f'下载失败: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)

