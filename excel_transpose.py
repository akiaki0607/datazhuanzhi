#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel合并单元格转置工具
将包含合并单元格的Excel表格转换为长格式数据
"""

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def analyze_excel_structure(file_path, sheet_name):
    """
    分析Excel文件结构，识别合并单元格
    """
    print(f"正在分析Excel文件: {file_path}")
    print(f"工作表名称: {sheet_name}")
    
    # 使用openpyxl读取工作簿
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 检查工作表是否存在
    if sheet_name not in wb.sheetnames:
        print(f"可用的工作表: {wb.sheetnames}")
        if wb.sheetnames:
            sheet_name = wb.sheetnames[0]  # 使用第一个工作表
            print(f"使用第一个工作表: {sheet_name}")
        else:
            raise ValueError("Excel文件中没有找到工作表")
    
    ws = wb[sheet_name]
    
    print(f"工作表尺寸: {ws.max_row} 行 x {ws.max_column} 列")
    
    # 分析合并单元格
    merged_ranges = list(ws.merged_cells.ranges)
    print(f"找到 {len(merged_ranges)} 个合并单元格区域:")
    
    for i, merged_range in enumerate(merged_ranges):
        print(f"  合并区域 {i+1}: {merged_range}")
        # 获取合并单元格的值
        top_left_cell = ws[merged_range.min_row][merged_range.min_col-1]
        print(f"    值: {top_left_cell.value}")
    
    return wb, ws, merged_ranges

def process_merged_cells_to_long_format(file_path, sheet_name=None):
    """
    将包含合并单元格的Excel数据转换为长格式
    """
    try:
        # 分析文件结构
        wb, ws, merged_ranges = analyze_excel_structure(file_path, sheet_name)
        
        # 读取数据（跳过可能的标题行）
        data_rows = []
        
        # 获取所有数据行
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):  # 跳过空行
                data_rows.append(list(row))
        
        if not data_rows:
            raise ValueError("Excel文件中没有找到数据")
        
        print(f"读取到 {len(data_rows)} 行数据")
        
        # 找到第一个数据行（通常是标题行）
        header_row = data_rows[0]
        print(f"标题行: {header_row}")
        
        # 处理合并单元格，创建品牌列
        processed_data = []
        
        # 分析第一行，识别品牌分组
        brand_groups = []
        current_brand = None
        brand_start_col = None
        
        for i, cell_value in enumerate(header_row):
            if cell_value is not None:
                # 检查是否是品牌名称（根据您的示例）
                if isinstance(cell_value, str) and ("科技" in cell_value or "客户" in cell_value or "竞品" in cell_value):
                    if current_brand is not None:
                        # 保存前一个品牌组
                        brand_groups.append({
                            'brand': current_brand,
                            'start_col': brand_start_col,
                            'end_col': i - 1
                        })
                    
                    current_brand = cell_value
                    brand_start_col = i
                elif current_brand is not None and i > brand_start_col:
                    # 这是品牌下的子列
                    continue
        
        # 添加最后一个品牌组
        if current_brand is not None:
            brand_groups.append({
                'brand': current_brand,
                'start_col': brand_start_col,
                'end_col': len(header_row) - 1
            })
        
        print(f"识别到品牌组: {brand_groups}")
        
        # 如果没有找到品牌分组，尝试其他方法
        if not brand_groups:
            # 尝试从第二行获取子列标题
            if len(data_rows) > 1:
                sub_headers = data_rows[1]
                print(f"子标题行: {sub_headers}")
                
                # 假设前两列是分组，后面的列是数据
                if len(sub_headers) >= 6:  # 根据您的示例，应该有6列数据
                    # 手动创建品牌分组
                    brand_groups = [
                        {'brand': '移山科技(客户)', 'start_col': 1, 'end_col': 4},
                        {'brand': '趣搜科技(核心竞品)', 'start_col': 5, 'end_col': 8}
                    ]
        
        # 创建新的数据结构
        new_data = []
        
        # 获取子列标题（通常是DeepSeek, Kimi, 元宝, 豆包等）
        sub_headers = []
        if len(data_rows) > 1:
            sub_headers = data_rows[1][1:]  # 跳过第一列（信源平台名称）
        
        # 如果没有子标题，使用默认的
        if not sub_headers or all(h is None for h in sub_headers):
            sub_headers = ['DeepSeek', 'Kimi', '元宝', '豆包']
        
        print(f"子列标题: {sub_headers}")
        
        # 处理数据行（从第三行开始，因为前两行是标题）
        data_start_row = 2
        
        for row_idx in range(data_start_row, len(data_rows)):
            row = data_rows[row_idx]
            if not any(cell is not None for cell in row):
                continue
                
            source_platform = row[0]  # 第一列是信源平台名称
            
            if source_platform is None:
                continue
            
            # 为每个品牌创建一行数据
            for brand_group in brand_groups:
                new_row = {
                    '信源平台名称': source_platform,
                    '品牌': brand_group['brand']
                }
                
                # 添加品牌对应的数据列
                start_col = brand_group['start_col']
                end_col = min(brand_group['end_col'], len(row) - 1)
                
                for i, sub_header in enumerate(sub_headers):
                    col_idx = start_col + i
                    if col_idx <= end_col and col_idx < len(row):
                        new_row[sub_header] = row[col_idx]
                    else:
                        new_row[sub_header] = None
                
                new_data.append(new_row)
        
        # 创建DataFrame
        df = pd.DataFrame(new_data)
        print(f"转换后的数据形状: {df.shape}")
        print("转换后的数据预览:")
        print(df.head(10))
        
        return df
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def save_processed_data(df, output_file):
    """
    保存处理后的数据到新的Excel文件
    """
    if df is None:
        print("没有数据需要保存")
        return
    
    try:
        # 创建新的Excel文件
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='转置后数据', index=False)
        
        print(f"数据已保存到: {output_file}")
        
        # 显示保存的文件信息
        file_size = os.path.getsize(output_file)
        print(f"文件大小: {file_size / 1024:.2f} KB")
        
    except Exception as e:
        print(f"保存文件时出现错误: {str(e)}")
        import traceback
        traceback.print_exc()

def main():
    """
    主函数
    """
    # 输入文件路径
    input_file = "2025916移山科技循环10次采集任务34词对外报表_待处理.xlsx"
    output_file = "2025916移山科技循环10次采集任务34词对外报表_转置后.xlsx"
    
    print("=" * 60)
    print("Excel合并单元格转置工具")
    print("=" * 60)
    
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 找不到输入文件 {input_file}")
        return
    
    # 处理数据
    df = process_merged_cells_to_long_format(input_file)
    
    if df is not None:
        # 保存结果
        save_processed_data(df, output_file)
        
        print("\n" + "=" * 60)
        print("处理完成!")
        print("=" * 60)
        print(f"输入文件: {input_file}")
        print(f"输出文件: {output_file}")
        print(f"转换后数据行数: {len(df)}")
        print(f"转换后数据列数: {len(df.columns)}")
    else:
        print("处理失败，请检查输入文件格式")

if __name__ == "__main__":
    main()

